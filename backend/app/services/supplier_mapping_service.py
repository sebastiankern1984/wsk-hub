"""Supplier column mapping CRUD + auto-detection service.

Handles per-supplier CSV/Excel column → Hub field mappings.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.supplier_column_mapping import SupplierColumnMapping
from app.services.hub_field_registry import (
    HUB_FIELD_MAP,
    HUB_FIELDS,
    STANDARD_COLUMN_MAP,
)

logger = logging.getLogger(__name__)


async def get_mappings(
    db: AsyncSession, supplier_id: int
) -> list[SupplierColumnMapping]:
    """Get all column mappings for a supplier."""
    result = await db.execute(
        select(SupplierColumnMapping)
        .where(SupplierColumnMapping.supplier_id == supplier_id)
        .order_by(SupplierColumnMapping.csv_column)
    )
    return list(result.scalars().all())


async def save_mappings(
    db: AsyncSession,
    supplier_id: int,
    mappings: list[dict[str, str]],
) -> list[SupplierColumnMapping]:
    """Replace all mappings for a supplier (bulk save).

    Args:
        mappings: list of {"csv_column": "...", "hub_field": "..."}
    """
    # Delete existing
    await db.execute(
        delete(SupplierColumnMapping).where(
            SupplierColumnMapping.supplier_id == supplier_id
        )
    )

    # Insert new
    new_mappings = []
    for m in mappings:
        csv_col = m["csv_column"].strip()
        hub_field = m["hub_field"].strip()

        if not csv_col or not hub_field:
            continue

        # Validate hub_field exists
        if hub_field not in HUB_FIELD_MAP:
            logger.warning("Unknown hub_field '%s' — skipping", hub_field)
            continue

        mapping = SupplierColumnMapping(
            supplier_id=supplier_id,
            csv_column=csv_col.lower(),
            hub_field=hub_field,
        )
        db.add(mapping)
        new_mappings.append(mapping)

    await db.flush()
    return new_mappings


async def get_column_map(
    db: AsyncSession, supplier_id: int
) -> dict[str, str]:
    """Get {csv_column: hub_field} dict for a supplier.

    Returns supplier-specific mapping if configured,
    otherwise falls back to STANDARD_COLUMN_MAP.
    """
    mappings = await get_mappings(db, supplier_id)
    if mappings:
        return {m.csv_column: m.hub_field for m in mappings}
    return dict(STANDARD_COLUMN_MAP)


def extract_headers_from_file(file_content: bytes, filename: str) -> list[str]:
    """Extract column headers from a CSV or Excel file.

    Handles our template format (category row + header row + type hints)
    by detecting type-hint rows and using the correct header row.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext in ("xlsx", "xls"):
        return _extract_excel_headers(file_content)
    else:
        return _extract_csv_headers(file_content)


def _extract_csv_headers(file_content: bytes) -> list[str]:
    """Extract headers from first line of a semicolon-delimited CSV."""
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = file_content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = file_content.decode("latin-1", errors="replace")

    first_line = text.split("\n")[0].strip()
    if not first_line:
        return []

    headers = [h.strip().strip('"') for h in first_line.split(";")]
    return [h for h in headers if h]


def _extract_excel_headers(file_content: bytes) -> list[str]:
    """Extract headers from an Excel file, detecting template format."""
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    # Detect our template format: row 3 (index 2) contains type hints
    TYPE_HINTS = {
        "text", "preis (z.b. 12,50)", "ganzzahl", "dezimalzahl",
        "datum (tt.mm.jjjj)", "ja/nein", "maß (zahl)", "gewicht (zahl)",
    }

    header_row_idx = 0  # default: first row is header

    if len(all_rows) >= 3:
        row3_values = [str(v).strip().lower() for v in all_rows[2] if v]
        if row3_values and all(v in TYPE_HINTS for v in row3_values):
            # Template format: row 1 = category, row 2 = headers
            header_row_idx = 1

    raw_headers = all_rows[header_row_idx]
    headers = [str(h).strip() for h in raw_headers if h]
    return [h for h in headers if h]


def _normalize(s: str) -> str:
    """Normalize a string for fuzzy matching: lowercase, strip, remove special chars."""
    s = s.strip().lower()
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def auto_detect_mappings(
    headers: list[str],
) -> list[dict[str, str | None]]:
    """Auto-detect hub field mappings for CSV headers.

    Priority:
    1. Exact match (label.lower() == header.lower())
    2. Normalized match (strip special chars, umlauts)
    3. Key match (header matches hub field key)
    4. Substring match (hub label contained in header or vice versa)
    """
    results = []

    # Pre-compute normalized labels and keys
    label_map = {f.label.lower(): f.key for f in HUB_FIELDS}
    normalized_label_map = {_normalize(f.label): f.key for f in HUB_FIELDS}
    key_set = {f.key for f in HUB_FIELDS}
    normalized_key_map = {_normalize(f.key): f.key for f in HUB_FIELDS}

    for header in headers:
        h_lower = header.strip().lower()
        h_norm = _normalize(header)

        # 1. Exact label match
        if h_lower in label_map:
            results.append({
                "csv_column": header.strip(),
                "hub_field": label_map[h_lower],
                "confidence": "exact",
            })
            continue

        # 2. Exact key match
        if h_lower.replace(" ", "_").replace("-", "_") in key_set:
            results.append({
                "csv_column": header.strip(),
                "hub_field": h_lower.replace(" ", "_").replace("-", "_"),
                "confidence": "exact",
            })
            continue

        # 3. Normalized label match
        if h_norm in normalized_label_map:
            results.append({
                "csv_column": header.strip(),
                "hub_field": normalized_label_map[h_norm],
                "confidence": "normalized",
            })
            continue

        # 4. Normalized key match
        if h_norm in normalized_key_map:
            results.append({
                "csv_column": header.strip(),
                "hub_field": normalized_key_map[h_norm],
                "confidence": "normalized",
            })
            continue

        # 5. Substring match (label in header or header in label)
        best_match = None
        for field in HUB_FIELDS:
            f_norm = _normalize(field.label)
            if len(f_norm) >= 3 and f_norm in h_norm:
                best_match = field.key
                break
            if len(h_norm) >= 3 and h_norm in f_norm:
                best_match = field.key
                break

        if best_match:
            results.append({
                "csv_column": header.strip(),
                "hub_field": best_match,
                "confidence": "fuzzy",
            })
        else:
            results.append({
                "csv_column": header.strip(),
                "hub_field": None,
                "confidence": "none",
            })

    return results
