"""Supplier CSV/Excel import service.

Parses semicolon-delimited CSV or Excel files from suppliers,
matches rows to existing products or creates new ones, and links
supplier_products entries.

Supports dynamic column mappings per supplier (from DB) or falls
back to standard Hub field labels / legacy FürSie format.

3-case matching logic:
  1. PZN → identity_map lookup → product exists? link.
     PZN in ABDA but not hub? auto-create from ABDA + link.
  2. GTIN Stück → product_eans lookup → product exists? link.
  3. NAN → products.nan lookup → product exists? link.
  4. No match → create skeleton product from supplier data + link.
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.identity_map import IdentityMap
from app.models.import_log import ImportLog
from app.models.product import Product
from app.models.product_ean import ProductEan
from app.models.product_hs_code import ProductHsCode
from app.models.supplier import Supplier
from app.models.supplier_product import SupplierProduct
from app.services.event_store import append_event
from app.services.hub_field_registry import STANDARD_COLUMN_MAP
from app.services.product_processing import (
    _get_or_create_manufacturer,
    _safe_float,
    _safe_int,
    process_abda_to_product,
)

logger = logging.getLogger(__name__)

# MwSt-Kennzeichen → VAT rate
VAT_MAP = {
    "red": 7.0,
    "vol": 19.0,
    "frei": 0.0,
}


def _parse_date(val: str | None) -> date | None:
    """Parse dd.mm.yyyy date strings."""
    if not val or not val.strip():
        return None
    val = val.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_price(val: str | None) -> float | None:
    """Parse price like '1,383' → 1.383."""
    if not val or not val.strip():
        return None
    try:
        return float(val.strip().replace(",", "."))
    except (ValueError, TypeError):
        return None


def _parse_bool(val: str | None) -> bool | None:
    """Parse 'ja'/'nein' → True/False."""
    if not val or not val.strip():
        return None
    v = val.strip().lower()
    if v in ("ja", "j", "yes", "1", "true"):
        return True
    if v in ("nein", "n", "no", "0", "false"):
        return False
    return None


def _strip_leading_zeros_ean(val: str | None) -> str | None:
    """Clean EAN: strip whitespace, keep leading zeros for 8/13/14 digit EANs."""
    if not val or not val.strip():
        return None
    v = val.strip()
    # Skip obvious placeholder EANs like "0000001170123"
    if v.lstrip("0") == "" or len(v.lstrip("0")) < 4:
        return None
    return v


def _convert_dimension_to_mm(value: str | None, unit: str | None) -> int | None:
    """Convert dimension value+unit to mm."""
    n = _safe_float(value)
    if n is None or n == 0:
        return None
    u = (unit or "mm").strip().lower()
    if u == "cm":
        n *= 10
    elif u == "m":
        n *= 1000
    return int(n)


def _convert_weight_to_g(value: str | None, unit: str | None) -> int | None:
    """Convert weight value+unit to grams."""
    n = _safe_float(value)
    if n is None or n == 0:
        return None
    u = (unit or "g").strip().lower()
    if u == "kg":
        n *= 1000
    return int(n)


# ── FürSie-specific column name → hub field mapping (legacy) ──────
# Kept for backwards compatibility when supplier has no DB mapping.
FUERSIE_COLUMN_MAP = {
    "bezugsweg": "bezugsweg",
    "nan": "nan",
    "einheit": "einheit",
    "lieferanten-artikelnummer": "supplier_sku",
    "gtin stück": "gtin_stueck",
    "gtin karton": "gtin_karton",
    "lieferanten wawi": "lieferanten_wawi",
    "lieferant rzf-nr": "lieferant_rzf_nr",
    "lieferantenname": "lieferantenname",
    "warengruppen nr.": "warengruppen_nr",
    "warengruppen bezeichnung": "warengruppe",
    "artikelbezeichnung 2": "name_short",
    "artikelbezeichnung 1": "name",
    "mengentext": "mengentext",
    "verpackungsinhalt": "verpackungsinhalt",
    "verpackungseinheit": "verpackungseinheit",
    "netto-gebinde-gewicht (in kg)": "netto_gewicht_kg",
    "ks": "ks",
    "ek-preis": "ek_preis",
    "mengencode": "mengencode",
    "gültig ab": "gueltig_ab",
    "gültig bis": "gueltig_bis",
    "staffel 1": "staffel_1",
    "staffel 2": "staffel_2",
    "staffel 3": "staffel_3",
    "staffel 4": "staffel_4",
    "staffel 5": "staffel_5",
    "staffel 6": "staffel_6",
    "staffel 7": "staffel_7",
    "staffel 8": "staffel_8",
    "staffel 9": "staffel_9",
    "eigenmarke": "eigenmarke",
    "restlaufzeit": "restlaufzeit",
    "mwst-kennzeichen": "mwst",
    "bemerkung": "bemerkung",
    "pzn": "pzn",
    "palettenfaktor in einheiten": "palettenfaktor",
    "stück breite": "stueck_breite",
    "stück breite einheit": "stueck_breite_einheit",
    "stück höhe": "stueck_hoehe",
    "stück höhe einheit": "stueck_hoehe_einheit",
    "stück tiefe": "stueck_tiefe",
    "stück tiefe einheit": "stueck_tiefe_einheit",
    "ve breite": "ve_breite",
    "ve breite einheit": "ve_breite_einheit",
    "ve höhe": "ve_hoehe",
    "ve höhe einheit": "ve_hoehe_einheit",
    "ve tiefe": "ve_tiefe",
    "ve tiefe einheit": "ve_tiefe_einheit",
    "stück gewicht": "stueck_gewicht",
    "stück gewicht einheit": "stueck_gewicht_einheit",
    "ve gewicht": "ve_gewicht",
    "ve gewicht einheit": "ve_gewicht_einheit",
    "gefahrengut": "gefahrengut",
    "displaykennzeichen": "displaykennzeichen",
    "zolltarifnummer": "zolltarifnummer",
    "jugendschutz": "jugendschutz",
    "altersgrenze": "altersgrenze",
    "eurelec": "eurelec",
    "ean ve typ": "ean_ve_typ",
    "pfand": "pfand",
    "uvp": "uvp",
}


def _build_header_map(
    fieldnames: list[str],
    column_map: dict[str, str] | None = None,
) -> dict[str, str]:
    """Build header → hub_field mapping.

    Priority:
    1. Explicit column_map (supplier-specific from DB)
    2. STANDARD_COLUMN_MAP (hub field labels)
    3. FUERSIE_COLUMN_MAP (legacy)
    4. Fallback: sanitized header as key
    """
    combined_map: dict[str, str] = {}
    # Start with standard, then overlay supplier-specific
    combined_map.update(STANDARD_COLUMN_MAP)
    combined_map.update(FUERSIE_COLUMN_MAP)
    if column_map:
        combined_map = column_map  # Supplier DB mapping takes full precedence

    header_map: dict[str, str] = {}
    for h in fieldnames:
        clean = h.strip().strip('"').lower()
        if clean in combined_map:
            header_map[h] = combined_map[clean]
        else:
            header_map[h] = clean.replace(" ", "_").replace("-", "_")

    return header_map


def _parse_csv_rows(
    file_content: bytes,
    column_map: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    """Parse semicolon-delimited CSV to list of dicts with normalised keys."""
    # Try utf-8-sig (BOM), then latin-1 fallback
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = file_content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = file_content.decode("latin-1", errors="replace")

    reader = csv.DictReader(io.StringIO(text), delimiter=";", quotechar='"')

    if not reader.fieldnames:
        return []

    header_map = _build_header_map(list(reader.fieldnames), column_map)

    rows: list[dict[str, str]] = []
    for raw_row in reader:
        row: dict[str, str] = {}
        for orig_header, value in raw_row.items():
            key = header_map.get(orig_header, orig_header)
            row[key] = (value or "").strip().strip('"')
        rows.append(row)

    return rows


def _parse_excel_rows(
    file_content: bytes,
    column_map: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    """Parse Excel file (xlsx/xls) to list of dicts with normalised keys.

    Reads first sheet, first row = headers, rest = data.
    Skips rows 2-3 if they look like type hints / example rows from our template
    (row 2 = category, row 3 = type hint like "Text", "Ganzzahl", etc.)
    """
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # Row 1 might be category row from our template, Row 2 = actual headers
    # We need to detect if this is our template format (has category + header + type hint rows)
    all_rows = list(rows_iter)
    wb.close()

    if not all_rows:
        return []

    # Detect template format: if row 3 contains only type hint words
    TYPE_HINTS = {"text", "preis (z.b. 12,50)", "ganzzahl", "dezimalzahl",
                  "datum (tt.mm.jjjj)", "ja/nein", "maß (zahl)", "gewicht (zahl)"}

    header_row_idx = 0  # default: first row is header
    data_start_idx = 1

    if len(all_rows) >= 4:
        # Check if row 3 (index 2) looks like type hints
        row3_values = [str(v).strip().lower() for v in all_rows[2] if v]
        if row3_values and all(v in TYPE_HINTS for v in row3_values):
            # This is our template format: row 1 = category, row 2 = headers, row 3 = types, row 4 = example
            header_row_idx = 1
            data_start_idx = 4  # Skip category, header, types, example

    # Extract headers
    raw_headers = all_rows[header_row_idx]
    fieldnames = [str(h).strip() if h else "" for h in raw_headers]

    header_map = _build_header_map(fieldnames, column_map)

    rows: list[dict[str, str]] = []
    for raw_row in all_rows[data_start_idx:]:
        if not any(raw_row):  # Skip empty rows
            continue
        row: dict[str, str] = {}
        for i, value in enumerate(raw_row):
            if i >= len(fieldnames) or not fieldnames[i]:
                continue
            orig_header = fieldnames[i]
            key = header_map.get(orig_header, orig_header)
            row[key] = str(value).strip() if value is not None else ""
        rows.append(row)

    return rows


def _parse_file_rows(
    file_content: bytes,
    filename: str,
    column_map: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    """Parse CSV or Excel file based on extension."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext in ("xlsx", "xls"):
        return _parse_excel_rows(file_content, column_map)
    else:
        return _parse_csv_rows(file_content, column_map)


async def _find_product_by_pzn(db: AsyncSession, pzn: str) -> Product | None:
    """Find product via identity_map PZN."""
    result = await db.execute(
        select(IdentityMap).where(
            IdentityMap.identity_type == "PZN",
            IdentityMap.identity_value == pzn,
        )
    )
    identity = result.scalar_one_or_none()
    if not identity:
        return None
    result = await db.execute(
        select(Product).where(Product.product_id == identity.product_id)
    )
    return result.scalar_one_or_none()


async def _find_product_by_ean(db: AsyncSession, ean: str) -> Product | None:
    """Find product via product_eans table."""
    result = await db.execute(
        select(ProductEan).where(ProductEan.ean_value == ean).limit(1)
    )
    pe = result.scalar_one_or_none()
    if not pe:
        return None
    return await db.get(Product, pe.product_id)


async def _find_product_by_nan(db: AsyncSession, nan: str) -> Product | None:
    """Find product via products.nan column."""
    result = await db.execute(
        select(Product).where(Product.nan == nan).limit(1)
    )
    return result.scalar_one_or_none()


async def _create_skeleton_product(
    db: AsyncSession,
    row: dict[str, str],
    user_id: str | None,
) -> Product:
    """Create a minimal product from supplier CSV data."""
    name = row.get("name") or row.get("name_short") or "Unbenannt"
    nan = row.get("nan") or None
    pzn = row.get("pzn") or None

    # VAT
    mwst_raw = (row.get("mwst") or "").lower()
    vat_rate = VAT_MAP.get(mwst_raw)

    product = Product(
        name=name,
        name_short=row.get("name_short") or None,
        nan=nan,
        pzn=pzn,
        warengruppe=row.get("warengruppe") or None,
        unit_size=row.get("mengentext") or None,
        size_value=_safe_float(row.get("verpackungsinhalt")),
        size_unit=row.get("verpackungseinheit") or None,
        units_per_ve=_safe_int(row.get("einheit")),
        piece_width_mm=_convert_dimension_to_mm(
            row.get("stueck_breite"), row.get("stueck_breite_einheit")
        ),
        piece_height_mm=_convert_dimension_to_mm(
            row.get("stueck_hoehe"), row.get("stueck_hoehe_einheit")
        ),
        piece_length_mm=_convert_dimension_to_mm(
            row.get("stueck_tiefe"), row.get("stueck_tiefe_einheit")
        ),
        case_width_mm=_convert_dimension_to_mm(
            row.get("ve_breite"), row.get("ve_breite_einheit")
        ),
        case_height_mm=_convert_dimension_to_mm(
            row.get("ve_hoehe"), row.get("ve_hoehe_einheit")
        ),
        case_length_mm=_convert_dimension_to_mm(
            row.get("ve_tiefe"), row.get("ve_tiefe_einheit")
        ),
        weight_piece_g=_convert_weight_to_g(
            row.get("stueck_gewicht"), row.get("stueck_gewicht_einheit")
        ),
        weight_ve_g=_convert_weight_to_g(
            row.get("ve_gewicht"), row.get("ve_gewicht_einheit")
        ),
        weight_g=_convert_weight_to_g(row.get("netto_gewicht_kg"), "kg"),
        ve_per_palette=_safe_int(row.get("palettenfaktor")),
        dg_flag=_parse_bool(row.get("gefahrengut")),
        shelf_life_days=_safe_int(row.get("restlaufzeit")),
        vat_rate=vat_rate,
        status="draft",
        version=1,
    )
    db.add(product)
    await db.flush()  # get product.id + product_id

    # Identity map entries
    if pzn:
        db.add(IdentityMap(
            identity_type="PZN",
            identity_value=pzn,
            product_id=product.product_id,
            source="supplier_import",
        ))
    if nan:
        db.add(IdentityMap(
            identity_type="NAN",
            identity_value=nan,
            product_id=product.product_id,
            source="supplier_import",
        ))

    # Event
    await append_event(
        db,
        event_type="ProductCreated",
        aggregate_type="product",
        aggregate_id=product.product_id,
        aggregate_version=product.version,
        payload={"name": product.name, "source": "supplier_import"},
        source="supplier_import",
        user_id=user_id,
    )

    await db.flush()
    return product


async def _add_eans_if_new(
    db: AsyncSession,
    product: Product,
    gtin_stueck: str | None,
    gtin_karton: str | None,
    source: str,
):
    """Add EANs to product if they don't already exist globally."""
    for ean_val, ean_type in [(gtin_stueck, "pack"), (gtin_karton, "ve")]:
        clean = _strip_leading_zeros_ean(ean_val)
        if not clean:
            continue
        # Check globally (unique constraint is on ean_type+ean_value where valid_to IS NULL)
        result = await db.execute(
            select(ProductEan).where(
                ProductEan.ean_type == ean_type,
                ProductEan.ean_value == clean,
                ProductEan.valid_to.is_(None),
            )
        )
        if not result.scalar_one_or_none():
            db.add(ProductEan(
                product_id=product.id,
                ean_type=ean_type,
                ean_value=clean,
                is_primary=(ean_type == "pack"),
                source=source,
            ))


async def _add_hs_code_if_new(
    db: AsyncSession,
    product: Product,
    zolltarifnummer: str | None,
    source: str = "supplier_import",
):
    """Add HS code if present and not already stored. Respects is_locked."""
    if not zolltarifnummer or not zolltarifnummer.strip():
        return
    hs = zolltarifnummer.strip()
    result = await db.execute(
        select(ProductHsCode).where(
            ProductHsCode.product_id == product.id,
            ProductHsCode.country == "DE",
        )
    )
    existing = result.scalar_one_or_none()
    if existing:
        # If locked, don't touch; if unlocked, update value
        if not existing.is_locked and existing.hs_code != hs:
            existing.hs_code = hs
            existing.source = source
    else:
        db.add(ProductHsCode(
            product_id=product.id,
            country="DE",
            hs_code=hs,
            source=source,
        ))


def _build_source_data(row: dict[str, str]) -> dict:
    """Collect non-mapped fields into a source_data dict for JSONB."""
    keys = [
        "bezugsweg", "lieferanten_wawi", "lieferant_rzf_nr", "ks",
        "mengencode", "eigenmarke", "displaykennzeichen", "jugendschutz",
        "altersgrenze", "eurelec", "ean_ve_typ", "pfand", "bemerkung",
        "warengruppen_nr",
    ]
    # Staffeln
    for i in range(1, 10):
        keys.append(f"staffel_{i}")

    data = {}
    for k in keys:
        v = row.get(k)
        if v:
            data[k] = v
    return data


async def _get_or_create_supplier(db: AsyncSession, name: str) -> Supplier:
    """Find supplier by name or create."""
    result = await db.execute(select(Supplier).where(Supplier.name == name))
    supplier = result.scalar_one_or_none()
    if not supplier:
        supplier = Supplier(name=name)
        db.add(supplier)
        await db.flush()
    return supplier


async def import_supplier_file(
    db: AsyncSession,
    import_log_id: int,
    file_content: bytes,
    filename: str = "import.csv",
    supplier_id: int | None = None,
    user_id: str | None = None,
):
    """Main import entry point. Runs in background task.

    Supports CSV and Excel files. If supplier_id is given, loads
    the supplier's column mapping from DB.
    """
    log = await db.get(ImportLog, import_log_id)
    if not log:
        logger.error("ImportLog %d not found", import_log_id)
        return

    log.status = "running"
    log.started_at = datetime.utcnow()
    await db.commit()

    try:
        # Load supplier-specific column mapping if available
        column_map: dict[str, str] | None = None
        if supplier_id:
            from app.services.supplier_mapping_service import get_column_map as _get_map
            db_map = await _get_map(db, supplier_id)
            # Only use if it's a real supplier mapping (not the standard fallback)
            from app.models.supplier_column_mapping import SupplierColumnMapping
            result = await db.execute(
                select(SupplierColumnMapping).where(
                    SupplierColumnMapping.supplier_id == supplier_id
                ).limit(1)
            )
            if result.scalar_one_or_none():
                column_map = db_map

        rows = _parse_file_rows(file_content, filename, column_map)
        log.total_rows = len(rows)
        await db.commit()

        imported = 0
        skipped = 0
        processed = 0

        # Group by supplier name to reuse supplier records
        supplier_cache: dict[str, Supplier] = {}

        for i, row in enumerate(rows):
            try:
                processed += 1

                # Skip rows without meaningful data
                article_name = row.get("name") or row.get("name_short")
                if not article_name or not article_name.strip():
                    skipped += 1
                    continue

                # ── Get/create supplier ──
                supplier_name = (row.get("lieferantenname") or "").strip()
                if not supplier_name:
                    skipped += 1
                    continue

                if supplier_name not in supplier_cache:
                    supplier_cache[supplier_name] = await _get_or_create_supplier(
                        db, supplier_name
                    )
                    # Commit supplier immediately so it survives row-level rollbacks
                    await db.commit()
                supplier = supplier_cache[supplier_name]

                # Use SAVEPOINT for per-row isolation
                async with db.begin_nested():
                    # ── Match product ──
                    product: Product | None = None
                    match_source = "none"

                    pzn = (row.get("pzn") or "").strip() or None
                    gtin_stueck = _strip_leading_zeros_ean(row.get("gtin_stueck"))
                    nan = (row.get("nan") or "").strip() or None

                    # Case 1: PZN match
                    if pzn:
                        product = await _find_product_by_pzn(db, pzn)
                        if product:
                            match_source = "pzn_hub"
                        else:
                            # Try ABDA auto-create
                            try:
                                product = await process_abda_to_product(
                                    db, pzn, user_id=user_id
                                )
                                match_source = "pzn_abda"
                            except ValueError:
                                pass  # Not in ABDA either

                    # Case 2: EAN match
                    if not product and gtin_stueck:
                        product = await _find_product_by_ean(db, gtin_stueck)
                        if product:
                            match_source = "ean"

                    # Case 3: NAN match
                    if not product and nan:
                        product = await _find_product_by_nan(db, nan)
                        if product:
                            match_source = "nan"

                    # Case 4: Skeleton
                    if not product:
                        product = await _create_skeleton_product(db, row, user_id)
                        match_source = "skeleton"

                    # ── Enrich product with supplier data (fill empty fields) ──
                    _enrich_product(product, row)

                    # ── Create supplier_product link ──
                    supplier_sku = (row.get("supplier_sku") or "").strip() or None

                    # Check if link already exists
                    existing_link = await db.execute(
                        select(SupplierProduct).where(
                            SupplierProduct.product_id == product.id,
                            SupplierProduct.supplier_id == supplier.id,
                        )
                    )
                    sp = existing_link.scalar_one_or_none()

                    ek_price = _parse_price(row.get("ek_preis"))
                    uvp = _parse_price(row.get("uvp"))
                    valid_from = _parse_date(row.get("gueltig_ab"))
                    valid_until = _parse_date(row.get("gueltig_bis"))
                    source_data = _build_source_data(row)

                    if sp:
                        # Update existing link
                        sp.supplier_sku = supplier_sku or sp.supplier_sku
                        if ek_price is not None:
                            sp.purchase_price = ek_price
                        if uvp is not None and uvp > 0:
                            sp.retail_price = uvp
                        sp.valid_from = valid_from or sp.valid_from
                        sp.valid_until = valid_until or sp.valid_until
                        sp.source_data = source_data
                    else:
                        db.add(SupplierProduct(
                            product_id=product.id,
                            supplier_id=supplier.id,
                            supplier_sku=supplier_sku,
                            purchase_price=ek_price,
                            retail_price=uvp if uvp and uvp > 0 else None,
                            valid_from=valid_from,
                            valid_until=valid_until,
                            source_data=source_data,
                        ))

                    # ── Add EANs ──
                    await _add_eans_if_new(
                        db, product,
                        row.get("gtin_stueck"),
                        row.get("gtin_karton"),
                        source=f"supplier:{supplier_name}",
                    )

                    # ── Add HS code ──
                    await _add_hs_code_if_new(db, product, row.get("zolltarifnummer"))

                imported += 1

                # Commit every 100 rows
                if processed % 100 == 0:
                    log.processed_rows = processed
                    log.imported_rows = imported
                    log.skipped_rows = skipped
                    await db.commit()
                    logger.info(
                        "Supplier import %d: %d/%d processed (%d imported, %d skipped)",
                        import_log_id, processed, len(rows), imported, skipped,
                    )

            except Exception as e:
                logger.error(
                    "Row %d error: %s — %s", i + 1, str(e), row.get("name", "?")
                )
                skipped += 1
                # SAVEPOINT rollback already happened, no need to rollback session
                continue

        log.processed_rows = processed
        log.imported_rows = imported
        log.skipped_rows = skipped
        log.status = "completed"
        log.completed_at = datetime.utcnow()
        await db.commit()

        logger.info(
            "Supplier import %d completed: %d imported, %d skipped of %d total",
            import_log_id, imported, skipped, len(rows),
        )

    except Exception as e:
        logger.exception("Supplier import %d failed: %s", import_log_id, str(e))
        await db.rollback()
        log = await db.get(ImportLog, import_log_id)
        if log:
            log.status = "failed"
            log.error_message = str(e)[:2000]
            log.completed_at = datetime.utcnow()
            await db.commit()


def _enrich_product(product: Product, row: dict[str, str]):
    """Fill empty product fields with supplier data (don't overwrite existing)."""
    locks = product.field_locks or {}

    def _set_if_empty(field: str, value):
        if value is not None and getattr(product, field, None) is None and not locks.get(field, False):
            setattr(product, field, value)

    _set_if_empty("nan", (row.get("nan") or "").strip() or None)
    _set_if_empty("name_short", row.get("name_short") or None)
    _set_if_empty("warengruppe", row.get("warengruppe") or None)
    _set_if_empty("unit_size", row.get("mengentext") or None)
    _set_if_empty("size_value", _safe_float(row.get("verpackungsinhalt")))
    _set_if_empty("size_unit", row.get("verpackungseinheit") or None)
    _set_if_empty("units_per_ve", _safe_int(row.get("einheit")))
    _set_if_empty("piece_width_mm", _convert_dimension_to_mm(
        row.get("stueck_breite"), row.get("stueck_breite_einheit")
    ))
    _set_if_empty("piece_height_mm", _convert_dimension_to_mm(
        row.get("stueck_hoehe"), row.get("stueck_hoehe_einheit")
    ))
    _set_if_empty("piece_length_mm", _convert_dimension_to_mm(
        row.get("stueck_tiefe"), row.get("stueck_tiefe_einheit")
    ))
    _set_if_empty("case_width_mm", _convert_dimension_to_mm(
        row.get("ve_breite"), row.get("ve_breite_einheit")
    ))
    _set_if_empty("case_height_mm", _convert_dimension_to_mm(
        row.get("ve_hoehe"), row.get("ve_hoehe_einheit")
    ))
    _set_if_empty("case_length_mm", _convert_dimension_to_mm(
        row.get("ve_tiefe"), row.get("ve_tiefe_einheit")
    ))
    _set_if_empty("weight_piece_g", _convert_weight_to_g(
        row.get("stueck_gewicht"), row.get("stueck_gewicht_einheit")
    ))
    _set_if_empty("weight_ve_g", _convert_weight_to_g(
        row.get("ve_gewicht"), row.get("ve_gewicht_einheit")
    ))
    _set_if_empty("weight_g", _convert_weight_to_g(row.get("netto_gewicht_kg"), "kg"))
    _set_if_empty("ve_per_palette", _safe_int(row.get("palettenfaktor")))
    _set_if_empty("dg_flag", _parse_bool(row.get("gefahrengut")))
    _set_if_empty("shelf_life_days", _safe_int(row.get("restlaufzeit")))
    mwst_raw = (row.get("mwst") or "").lower()
    _set_if_empty("vat_rate", VAT_MAP.get(mwst_raw))
    _set_if_empty("hs_code", (row.get("zolltarifnummer") or "").strip() or None)
