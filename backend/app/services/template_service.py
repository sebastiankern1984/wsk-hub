"""Template generation for supplier import files.

Generates downloadable CSV and Excel templates with all Hub fields
or supplier-specific fields (based on their column mapping).
"""
from __future__ import annotations

import csv
import io
import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.hub_field_registry import HUB_FIELD_MAP, HUB_FIELDS, HubField

logger = logging.getLogger(__name__)

# Category display names (German)
CATEGORY_LABELS = {
    "identifikation": "Identifikation",
    "artikeldaten": "Artikeldaten",
    "preis": "Preise",
    "masse": "Maße & Gewicht",
    "logistik": "Logistik",
    "sonstiges": "Sonstiges",
}

# Category colors for Excel headers
CATEGORY_COLORS = {
    "identifikation": "4472C4",
    "artikeldaten": "548235",
    "preis": "BF8F00",
    "masse": "7030A0",
    "logistik": "C00000",
    "sonstiges": "808080",
}


def _get_fields_for_mappings(
    mappings: list[dict[str, str]] | None,
) -> list[tuple[str, HubField]]:
    """Get ordered list of (csv_header, hub_field) for template generation.

    If mappings provided: use supplier's CSV column names → hub field.
    Otherwise: use all Hub fields with labels as headers.
    """
    if mappings:
        result = []
        for m in mappings:
            hf = HUB_FIELD_MAP.get(m["hub_field"])
            if hf:
                result.append((m["csv_column"], hf))
        return result

    # Standard: all fields, labels as headers
    return [(f.label, f) for f in HUB_FIELDS]


# ── CSV Template ──────────────────────────────────────────────────


def generate_template_csv(
    mappings: list[dict[str, str]] | None = None,
) -> bytes:
    """Generate a CSV template (semicolon-delimited, UTF-8 BOM).

    Returns bytes ready for download.
    """
    fields = _get_fields_for_mappings(mappings)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # Header row
    writer.writerow([csv_header for csv_header, _ in fields])

    # Example row
    writer.writerow([hf.example for _, hf in fields])

    # UTF-8 BOM + content
    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


# ── Excel Template ────────────────────────────────────────────────


def generate_template_xlsx(
    mappings: list[dict[str, str]] | None = None,
) -> bytes:
    """Generate an Excel template with formatting and data type hints.

    Returns bytes (xlsx) ready for download.
    """
    fields = _get_fields_for_mappings(mappings)

    wb = Workbook()
    ws = wb.active
    ws.title = "Import-Vorlage"

    # ── Header row (Row 1): Category group ──
    # ── Header row (Row 2): Column names ──
    # ── Type hint row (Row 3): Expected format ──
    # ── Example row (Row 4): Example values ──

    header_font = Font(bold=True, color="FFFFFF", size=11)
    type_font = Font(italic=True, color="666666", size=9)
    example_fill = PatternFill("solid", fgColor="F2F2F2")

    # Type descriptions
    TYPE_HINTS = {
        "str": "Text",
        "price": "Preis (z.B. 12,50)",
        "int": "Ganzzahl",
        "float": "Dezimalzahl",
        "date": "Datum (TT.MM.JJJJ)",
        "bool": "ja/nein",
        "dimension": "Maß (Zahl)",
        "weight": "Gewicht (Zahl)",
    }

    current_category = None
    col = 1

    for csv_header, hf in fields:
        # Row 1: Category (merged later if needed)
        cat_label = CATEGORY_LABELS.get(hf.category, hf.category)
        cat_color = CATEGORY_COLORS.get(hf.category, "808080")
        cat_fill = PatternFill("solid", fgColor=cat_color)

        cell_cat = ws.cell(row=1, column=col, value=cat_label)
        cell_cat.font = Font(bold=True, color="FFFFFF", size=9)
        cell_cat.fill = cat_fill

        # Row 2: Column header
        cell_header = ws.cell(row=2, column=col, value=csv_header)
        cell_header.font = header_font
        cell_header.fill = PatternFill("solid", fgColor="333333")
        cell_header.alignment = Alignment(wrap_text=True)

        # Row 3: Type hint
        type_hint = TYPE_HINTS.get(hf.field_type, "Text")
        cell_type = ws.cell(row=3, column=col, value=type_hint)
        cell_type.font = type_font

        # Row 4: Example value
        cell_example = ws.cell(row=4, column=col, value=hf.example)
        cell_example.fill = example_fill

        # Column width (auto)
        max_len = max(len(csv_header), len(hf.example), len(type_hint), 10)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)

        col += 1

    # Freeze header rows
    ws.freeze_panes = "A5"

    # ── Info sheet ──
    ws_info = wb.create_sheet("Info")
    ws_info.cell(row=1, column=1, value="WSK Hub Import-Vorlage").font = Font(bold=True, size=14)
    ws_info.cell(row=3, column=1, value="Hinweise:")
    ws_info.cell(row=4, column=1, value="• Zeile 1: Kategorie")
    ws_info.cell(row=5, column=1, value="• Zeile 2: Spaltenname (nicht ändern!)")
    ws_info.cell(row=6, column=1, value="• Zeile 3: Erwartetes Format")
    ws_info.cell(row=7, column=1, value="• Zeile 4: Beispielwerte (löschen vor Import)")
    ws_info.cell(row=8, column=1, value="• Ab Zeile 5: Ihre Daten einfügen")
    ws_info.cell(row=10, column=1, value="Trennzeichen CSV: Semikolon (;)")
    ws_info.cell(row=11, column=1, value="Encoding CSV: UTF-8")
    ws_info.cell(row=12, column=1, value="Datumsformat: TT.MM.JJJJ")
    ws_info.cell(row=13, column=1, value="Preise: Komma als Dezimaltrenner (z.B. 12,50)")
    ws_info.column_dimensions["A"].width = 60

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
