from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_import_pool
from app.models.abda import AbdaImportLog, AbdaPacApo, AbdaPriceHistory

logger = logging.getLogger(__name__)

# Build set of valid DB column names from the model
_VALID_COLUMNS = {c.name for c in AbdaPacApo.__table__.columns} - {"_updated_at"}

# Column mapping: Excel header → DB column
CURATED_COLUMN_MAP = {
    "PZN": "pzn",
    "EAN": "gtin",
    "Apo_Ek (EUR)": "apo_ek",
    "Name": "langname",
    "Hersteller": "hersteller_name",
    "Packungsgröße": "packungsgroesse",
    "Normgröße": "normgroesse",
    "Breite (mm)": "breite",
    "Höhe (mm)": "hoehe",
    "Länge (mm)": "laenge",
    "Gewicht (g)": "gewicht",
    "Arzneimittel": "arzneimittel",
    "MwSt": "mwst",
    "Apothekenpflicht": "apopflicht",
    "Verkehrsstatus": "verkehrsstatus",
    "Vertriebsstatus": "vertriebsstatus",
    "Preis gültig ab": "gdat_preise",
}

BATCH_SIZE = 5000


async def import_abda_excel(db: AsyncSession, import_log_id: int, file_path: str) -> None:
    """Import ABDA Excel file into abda_pac_apo table."""
    log = await db.get(AbdaImportLog, import_log_id)
    if not log:
        return

    log.status = "running"
    log.started_at = datetime.utcnow()
    await db.commit()

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel-Datei hat kein aktives Arbeitsblatt")

        rows = ws.iter_rows()
        header_row = next(rows)
        headers = [str(cell.value).strip() if cell.value else "" for cell in header_row]

        # Build column mapping: excel_col_index → db_column_name
        col_map: dict[int, str] = {}
        for idx, header in enumerate(headers):
            if header in CURATED_COLUMN_MAP:
                col_map[idx] = CURATED_COLUMN_MAP[header]
            elif header.lower() in _VALID_COLUMNS and header.lower() != "pzn":
                # Raw fields: use lowercase header as DB column (only if column exists)
                col_map[idx] = header.lower()

        # PZN column index
        pzn_idx = None
        for idx, header in enumerate(headers):
            if header == "PZN":
                pzn_idx = idx
                break

        if pzn_idx is None:
            raise ValueError("PZN-Spalte nicht in Excel gefunden")

        total = 0
        inserts = 0
        updates = 0
        batch: list[dict[str, str | None]] = []

        pool = await get_import_pool()

        for row in rows:
            pzn_val = str(row[pzn_idx].value).strip() if row[pzn_idx].value else None
            if not pzn_val:
                continue

            record: dict[str, str | None] = {"pzn": pzn_val}
            for col_idx, db_col in col_map.items():
                if db_col == "pzn":
                    continue
                cell_val = row[col_idx].value if col_idx < len(row) else None
                record[db_col] = str(cell_val).strip() if cell_val is not None else None

            batch.append(record)
            total += 1

            if len(batch) >= BATCH_SIZE:
                ins, upd = await _process_batch(pool, batch, db, import_log_id)
                inserts += ins
                updates += upd
                batch = []

                # Update progress
                log.record_count_total = total
                log.record_count_insert = inserts
                log.record_count_update = updates
                await db.commit()

        # Process remaining batch
        if batch:
            ins, upd = await _process_batch(pool, batch, db, import_log_id)
            inserts += ins
            updates += upd

        wb.close()

        log.record_count_total = total
        log.record_count_insert = inserts
        log.record_count_update = updates
        log.status = "completed"
        log.completed_at = datetime.utcnow()
        await db.commit()

        logger.info(f"ABDA import completed: {total} total, {inserts} inserts, {updates} updates")

    except Exception as e:
        logger.exception(f"ABDA import failed: {e}")
        log.status = "failed"
        log.error_message = str(e)[:2000]
        log.completed_at = datetime.utcnow()
        await db.commit()


async def _process_batch(
    pool,
    batch: list[dict[str, str | None]],
    db: AsyncSession,
    import_log_id: int,
) -> tuple[int, int]:
    """Process a batch of records. Returns (inserts, updates)."""
    if not batch:
        return 0, 0

    inserts = 0
    updates = 0

    async with pool.acquire() as conn:
        # Get existing PZNs and their apo_ek for price tracking
        pzns = [r["pzn"] for r in batch]
        existing = {}
        rows = await conn.fetch(
            "SELECT pzn, apo_ek FROM abda_pac_apo WHERE pzn = ANY($1)",
            pzns,
        )
        for row in rows:
            existing[row["pzn"]] = row["apo_ek"]

        for record in batch:
            pzn = record["pzn"]
            is_update = pzn in existing

            # Build columns and values for upsert
            cols = [k for k in record.keys() if record[k] is not None]
            vals = [record[k] for k in cols]
            cols.append("_updated_at")
            vals.append(datetime.utcnow())

            placeholders = ", ".join(f"${i+1}" for i in range(len(vals)))
            col_names = ", ".join(cols)
            update_set = ", ".join(f"{c} = EXCLUDED.{c}" for c in cols if c != "pzn")

            await conn.execute(
                f"INSERT INTO abda_pac_apo ({col_names}) VALUES ({placeholders}) "
                f"ON CONFLICT (pzn) DO UPDATE SET {update_set}",
                *vals,
            )

            if is_update:
                updates += 1
                # Track price changes
                old_price = existing.get(pzn)
                new_price = record.get("apo_ek")
                if old_price != new_price and (old_price or new_price):
                    await conn.execute(
                        "INSERT INTO abda_price_history (pzn, old_apo_ek, new_apo_ek, change_source) "
                        "VALUES ($1, $2, $3, $4)",
                        pzn,
                        old_price,
                        new_price,
                        f"import_{import_log_id}",
                    )
            else:
                inserts += 1

    return inserts, updates
